import openpyxl

# Открываем файл Excel
workbook = openpyxl.load_workbook('04_09_2023_г_Сводная_таблица_ИП_Чертыков_56_актов_5.xlsx')

# Выбираем нужный лист
sheet = workbook['Основной список']

# Создаем новый лист для вставки строк
new_sheet = workbook.create_sheet('Новый список', 1)

# Проходимся по каждой строке в таблице
for row in sheet.iter_rows(min_row=1, min_col=1, max_col=36):
    # Получаем значение ячейки, которая содержит количество пустых строк для добавления
    cell = row[3]  # Например, пятый столбец (индекс 4)

    # Получаем количество пустых строк
    num_empty_rows = cell.value if cell.value is not None else 0

    # Копируем текущую строку в новый лист
    new_sheet.append([cell.value for cell in row])

    # Вставляем пустые строки
    for _ in range(int(num_empty_rows)):
        new_sheet.append([None] * 14)

# Удаляем старый лист
workbook.remove(sheet)

# Переименовываем новый лист в оригинальное имя
new_sheet.title = 'Основной список'

# Сохраняем изменения
workbook.save('export.xlsx')