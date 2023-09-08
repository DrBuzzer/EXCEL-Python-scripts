import openpyxl

# Открываем файл Excel
workbook = openpyxl.load_workbook('Ваш_файл.xlsx')

# Выбираем нужный лист
sheet = workbook['Имя_вашего_листа']

# Создаем новый лист для обновленных данных
updated_sheet = workbook.create_sheet('Обновленные данные')

# Проходимся по каждой строке в листе
for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
    for cell in row:
        # Получаем значение ячейки
        value = cell.value

        # Разделяем значение на отдельные строки
        rows = str(value).split('\n') if value else ['']

        # Устанавливаем значения разделенных строк в новые ячейки и строки
        for i, row_value in enumerate(rows):
            updated_sheet.cell(row=cell.row + i, column=cell.column).value = row_value

# Сохраняем изменения
workbook.save('export.xlsx')